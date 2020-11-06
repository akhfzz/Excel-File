from openpyxl import load_workbook,Workbook
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

#define = Workbook() #membuka sistem
#define.save("Arazor.xlsx") #membuat file dengan nama berikut
def Convert(value):
	if isinstance(value,bool):#memeriksa bahwa objek merupakan instance dari suatu kelas
		return value #mengeluarkan value
	try:#mencoba dalam mengkonversikan parameter value menjadi integer
		return int(value)
	except ValueError:#ternyata value eror dan menjadikannya sebagai return value
		try:
			return float(value)#return value
		except ValueError:
			return value 

Open = Workbook()#membuka sistem ekstensi pada excel
maker = Open.active#mengaktifkan
maker.title = "Data Mahasiswa"#membuat nama sheet

maker.append(["NIM", "Fname", "Lname", "Jurusan", "IP1", "IP2", "IP3"])#membuat atribut
Things = [{"NIM":192102002, "Fname": "Akhmad", "Lname": "Faizal", "Jurusan":"S-1 Informatika", "IP1":"3,56", "IP2": "3,75", "IP3": "3,89"},
			{"NIM":192102001, "Fname": "Agnes","Lname": "Silvina", "Jurusan":"S-1 Informatika", "IP1":"3,16", "IP2": "3,25", "IP3": "3,69"},
			{"NIM":192102003, "Fname": "Alfabeth","Lname": "Danuningrat", "Jurusan":"S-1 Informatika", "IP1":"3,38", "IP2": "3,05", "IP3": "3,39"},
			{"NIM":192102124, "Fname": "Dedi","Lname": "Rizaldi", "Jurusan":"S-1 Akuntansi", "IP1":"3,50", "IP2": "3,78", "IP3": "3,29"}]
			#Things di gunakan untuk membuat atribut menjadi key serta memiliki value
for mahasiswa in Things: #pengaplikasiannya
	maker.append(list(map(Convert,mahasiswa.values())))#membuat list serta memetakan sebuah nilai
#	print(mahasiswa.values())
#	print(list(map(Convert,mahasiswa.values())))
#last_cell = maker.cell(row = maker.max_row, column= maker.max_column).coordinate# #membuat table dengan kordinasi dengan software jumlah kolom dan baris
#table_name = Table(displayName= "DataMhs", ref ="A1:{}".format(last_cell))#membuat nama table,serta dari mana table di mulai
#Origin = TableStyleInfo(name= "TableStyleMedium6", showRowStripes= True) #membuat table style
#table_name.TableStyleInfo = Origin #menghubungkan kordinasi pembuatan table dengan table style
#maker.add_table(table_name) #menjadikan table di implementasikan
#Open.save("Arazor.xlsx")#simpan pada file perubahan


#Average = IP1 + IP2 + IP3/3
Open = load_workbook(filename= "Arazor.xlsx") #melakukan pemetaan table
maker = Open["Data Mahasiswa"] #membuka sheet pada file tersebut
maker["H1"] = "Average" #membuat atribut baru pada sell ke sekian

for r in range(2, maker.max_row + 1): #perulangan pada tabel 2 untuk pengisian nilai tabel
	ip1 = maker.cell(row=r, column=5).value
	ip2 = maker.cell(row=r, column=6).value
	ip3 = maker.cell(row=r, column=7).value
	#var ip1,ip2,ip3 menentukan kolom mana yang akan di isi value nya
	avrg = round(ip1+ip2+ip3/3,2) #perhitungan serta angka menyusun indeks yang di tampilkan
	maker.cell(row= r, column=8).value= avrg# menetukan perhitungan pada sell hasil
Open.save("Arazor.xlsx") #simpan


